"""Excel report generation for the Cash Flow Projection Tool."""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Color / style constants
# ---------------------------------------------------------------------------
_NAVY = "1F3864"
_WHITE = "FFFFFF"
_LIGHT_BLUE = "D6E4F0"
_VERY_LIGHT_BLUE = "EBF5FB"
_ALT_GREY = "F2F2F2"
_RED_FONT = "C0392B"
_GREEN_FILL = "D5E8D4"
_RED_FILL = "F8CECC"

_THIN_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_FMT_DOLLAR = '$#,##0.00'
_FMT_PCT = '0.00%'
_FMT_NUMBER = '0.00'


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_cell_style(cell, style_name, negative=False):
    """Apply a named style to a cell."""
    base_font = Font(name="Calibri", size=11)

    if style_name == "title":
        cell.font = Font(name="Calibri", size=14, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    elif style_name == "subtitle":
        cell.font = Font(name="Calibri", size=10, italic=True, color="444444")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    elif style_name == "col_header":
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "row_label":
        cell.font = Font(name="Calibri", size=11, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "row_label_bold":
        cell.font = Font(name="Calibri", size=11, bold=True, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "data":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "data_section":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, bold=True, color=color)
        cell.fill = _fill(_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "data_margin":
        color = _RED_FONT if negative else "555555"
        cell.font = Font(name="Calibri", size=10, italic=True, color=color)
        cell.fill = _fill(_VERY_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "row_label_section":
        cell.font = Font(name="Calibri", size=11, bold=True, color="222222")
        cell.fill = _fill(_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "row_label_margin":
        cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
        cell.fill = _fill(_VERY_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        cell.border = _THIN_BORDER

    elif style_name == "summary_header":
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "summary_label":
        cell.font = Font(name="Calibri", size=11, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "summary_value":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, bold=True, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    else:
        cell.font = base_font


def _set_column_widths(ws, num_years):
    """Set col A = 32, cols B+ = 16."""
    ws.column_dimensions["A"].width = 32
    for col in range(2, num_years + 3):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _write_row(ws, row_num, label, values, label_style, data_style, num_format,
               alt=False, neg_check=True):
    """Write one data row: label in col A, values in cols B+."""
    label_cell = ws.cell(row=row_num, column=1, value=label)
    _apply_cell_style(label_cell, label_style)
    if alt and label_style == "row_label":
        label_cell.fill = _fill(_ALT_GREY)

    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row_num, column=col_offset + 2)
        if val is None:
            cell.value = "N/A"
            _apply_cell_style(cell, data_style)
        else:
            cell.value = val
            negative = neg_check and isinstance(val, (int, float)) and val < 0
            _apply_cell_style(cell, data_style, negative=negative)
            cell.number_format = num_format
        if alt and data_style == "data":
            cell.fill = _fill(_ALT_GREY)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_cashflow_sheet(wb, project_data, results):
    ws = wb.active
    ws.title = "Cash Flow Projection"

    company = project_data["company_name"]
    wacc = project_data["wacc"]
    initial_investment = project_data["initial_investment"]
    num_years = project_data["num_years"]
    year_metrics = results["year_metrics"]

    _set_column_widths(ws, num_years)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 22

    # Row 1: Title (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_years + 1)
    title_cell = ws.cell(row=1, column=1, value=f"{company} – Cash Flow Projection")
    _apply_cell_style(title_cell, "title")

    # Row 2: Subtitle
    subtitle = (
        f"WACC: {wacc*100:.2f}%  |  "
        f"Initial Investment: ${initial_investment:,.2f}  |  "
        f"Years: {num_years}"
    )
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    _apply_cell_style(sub_cell, "subtitle")

    # Row 3: blank

    # Row 4: Column headers
    header_label = ws.cell(row=4, column=1, value="Metric")
    _apply_cell_style(header_label, "col_header")
    for i in range(1, num_years + 1):
        cell = ws.cell(row=4, column=i + 1, value=f"Year {i}")
        _apply_cell_style(cell, "col_header")

    # Helper to extract a field across all years
    def field_values(key):
        return [m[key] for m in year_metrics]

    # Income statement rows 5–18
    rows_def = [
        # (label, field_key, label_style, data_style, num_format)
        ("Revenue",                "revenue",          "row_label_bold",   "data_section", _FMT_DOLLAR),
        ("COGS",                   "cogs",             "row_label",        "data",         _FMT_DOLLAR),
        ("Gross Profit",           "gross_profit",     "row_label_section","data_section", _FMT_DOLLAR),
        ("  Gross Margin %",       "gross_margin_pct", "row_label_margin", "data_margin",  _FMT_PCT),
        ("Operating Expenses",     "opex",             "row_label",        "data",         _FMT_DOLLAR),
        ("EBITDA",                 "ebitda",           "row_label_section","data_section", _FMT_DOLLAR),
        ("  EBITDA Margin %",      "ebitda_margin_pct","row_label_margin", "data_margin",  _FMT_PCT),
        ("D&A",                    "da",               "row_label",        "data",         _FMT_DOLLAR),
        ("EBIT",                   "ebit",             "row_label_bold",   "data",         _FMT_DOLLAR),
        ("Interest Expense",       "interest",         "row_label",        "data",         _FMT_DOLLAR),
        ("EBT",                    "ebt",              "row_label_bold",   "data",         _FMT_DOLLAR),
        ("Tax",                    "tax",              "row_label",        "data",         _FMT_DOLLAR),
        ("Net Income",             "net_income",       "row_label_section","data_section", _FMT_DOLLAR),
        ("  Net Margin %",         "net_margin_pct",   "row_label_margin", "data_margin",  _FMT_PCT),
    ]

    alt = False
    for row_offset, (label, key, lbl_style, dat_style, fmt) in enumerate(rows_def):
        row_num = 5 + row_offset
        vals = field_values(key)
        use_alt = alt and lbl_style == "row_label"
        _write_row(ws, row_num, label, vals, lbl_style, dat_style, fmt, alt=use_alt)
        if lbl_style == "row_label":
            alt = not alt

    # Row 19: blank separator
    ws.row_dimensions[19].height = 8

    # FCF rows 20–26
    fcf_rows = [
        ("D&A Add-back",              "da_addback",                "row_label",        "data",         _FMT_DOLLAR),
        ("CapEx",                     "capex",                     "row_label",        "data",         _FMT_DOLLAR),
        ("Δ Working Capital",         "delta_wc",                  "row_label",        "data",         _FMT_DOLLAR),
        ("Free Cash Flow (FCF)",      "fcf",                       "row_label_section","data_section", _FMT_DOLLAR),
        ("Discounted FCF",            "discounted_fcf",            "row_label",        "data",         _FMT_DOLLAR),
        ("Cumulative FCF",            "cumulative_fcf",            "row_label",        "data",         _FMT_DOLLAR),
        ("Cumulative Discounted FCF", "cumulative_discounted_fcf", "row_label",        "data",         _FMT_DOLLAR),
    ]

    alt = False
    for row_offset, (label, key, lbl_style, dat_style, fmt) in enumerate(fcf_rows):
        row_num = 20 + row_offset
        vals = field_values(key)
        use_alt = alt and lbl_style == "row_label"
        _write_row(ws, row_num, label, vals, lbl_style, dat_style, fmt, alt=use_alt)
        if lbl_style == "row_label":
            alt = not alt

    # Freeze panes: keep header and metric label column visible
    ws.freeze_panes = "B5"


def _build_summary_sheet(wb, project_data, results):
    ws = wb.create_sheet(title="Financial Summary")

    company = project_data["company_name"]
    wacc = project_data["wacc"]
    initial_investment = project_data["initial_investment"]
    num_years = project_data["num_years"]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[1].height = 26

    # Row 1: Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value=f"Financial Summary – {company}")
    _apply_cell_style(title_cell, "title")

    # Row 2: generation date
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {date.today().strftime('%B %d, %Y')}")
    _apply_cell_style(date_cell, "subtitle")

    # Row 4: headers
    ws.cell(row=4, column=1, value="Indicator")
    ws.cell(row=4, column=2, value="Value")
    _apply_cell_style(ws.cell(row=4, column=1), "summary_header")
    _apply_cell_style(ws.cell(row=4, column=2), "summary_header")

    def _summary_row(row, label, value, num_format=None, fill_hex=None):
        lbl = ws.cell(row=row, column=1, value=label)
        _apply_cell_style(lbl, "summary_label")

        negative = isinstance(value, (int, float)) and value < 0
        val_cell = ws.cell(row=row, column=2)
        if value is None:
            val_cell.value = "Not within projection"
            _apply_cell_style(val_cell, "summary_value")
        elif isinstance(value, str):
            val_cell.value = value
            _apply_cell_style(val_cell, "summary_value")
        else:
            val_cell.value = value
            _apply_cell_style(val_cell, "summary_value", negative=negative)
            if num_format:
                val_cell.number_format = num_format

        if fill_hex:
            val_cell.fill = _fill(fill_hex)
            lbl.fill = _fill(fill_hex)

    # Row 5: NPV
    npv = results["npv"]
    npv_fill = _GREEN_FILL if npv >= 0 else _RED_FILL
    _summary_row(5, "Net Present Value (NPV)", npv, _FMT_DOLLAR, npv_fill)

    # Row 6: IRR
    irr = results["irr"]
    if irr is None:
        _summary_row(6, "Internal Rate of Return (IRR)", "N/A")
    else:
        irr_fill = _GREEN_FILL if irr > wacc else _RED_FILL
        _summary_row(6, "Internal Rate of Return (IRR)", irr, _FMT_PCT, irr_fill)

    # Row 7: Simple Payback
    sp = results["simple_payback"]
    if sp is None:
        _summary_row(7, "Simple Payback Period", None)
    else:
        lbl = ws.cell(row=7, column=1, value="Simple Payback Period (years)")
        _apply_cell_style(lbl, "summary_label")
        val = ws.cell(row=7, column=2, value=sp)
        _apply_cell_style(val, "summary_value")
        val.number_format = _FMT_NUMBER

    # Row 8: Discounted Payback
    dp = results["discounted_payback"]
    if dp is None:
        _summary_row(8, "Discounted Payback Period", None)
    else:
        lbl = ws.cell(row=8, column=1, value="Discounted Payback Period (years)")
        _apply_cell_style(lbl, "summary_label")
        val = ws.cell(row=8, column=2, value=dp)
        _apply_cell_style(val, "summary_value")
        val.number_format = _FMT_NUMBER

    # Row 9: blank

    # Rows 10-12: Average margins
    _summary_row(10, "Avg. Gross Margin", results["avg_gross_margin"], _FMT_PCT)
    _summary_row(11, "Avg. EBITDA Margin", results["avg_ebitda_margin"], _FMT_PCT)
    _summary_row(12, "Avg. Net Margin", results["avg_net_margin"], _FMT_PCT)

    # Row 13: blank

    # Rows 14-16: Inputs reference
    _summary_row(14, "WACC / Discount Rate", wacc, _FMT_PCT)
    _summary_row(15, "Initial Investment", initial_investment, _FMT_DOLLAR)
    _summary_row(16, "Projection Years", str(num_years))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_excel(project_data, results, filename):
    """Create workbook, build sheets, and save."""
    wb = Workbook()
    _build_cashflow_sheet(wb, project_data, results)
    _build_summary_sheet(wb, project_data, results)
    wb.save(filename)
