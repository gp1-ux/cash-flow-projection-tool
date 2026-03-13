"""Generación de reportes Excel para la Herramienta de Proyección de Flujo de Caja (Español)."""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constantes de color / estilo
# ---------------------------------------------------------------------------
_NAVY = "1F3864"
_WHITE = "FFFFFF"
_LIGHT_BLUE = "D6E4F0"
_VERY_LIGHT_BLUE = "EBF5FB"
_ALT_GREY = "F2F2F2"
_RED_FONT = "C0392B"
_GREEN_FILL = "D5E8D4"
_RED_FILL = "F8CECC"

_THIN_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_FMT_DOLLAR = '$#,##0.00'
_FMT_PCT = '0.00%'
_FMT_NUMBER = '0.00'


# ---------------------------------------------------------------------------
# Ayudantes internos
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_cell_style(cell, style_name, negative=False):
    """Aplica un estilo nombrado a una celda."""
    if style_name == "title":
        cell.font = Font(name="Calibri", size=14, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    elif style_name == "subtitle":
        cell.font = Font(name="Calibri", size=10, italic=True, color="444444")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    elif style_name == "col_header":
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "row_label":
        cell.font = Font(name="Calibri", size=11, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "row_label_bold":
        cell.font = Font(name="Calibri", size=11, bold=True, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "data":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "data_section":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, bold=True, color=color)
        cell.fill = _fill(_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "data_margin":
        color = _RED_FONT if negative else "555555"
        cell.font = Font(name="Calibri", size=10, italic=True, color=color)
        cell.fill = _fill(_VERY_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "row_label_section":
        cell.font = Font(name="Calibri", size=11, bold=True, color="222222")
        cell.fill = _fill(_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "row_label_margin":
        cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
        cell.fill = _fill(_VERY_LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        cell.border = _THIN_BORDER

    elif style_name == "summary_header":
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    elif style_name == "summary_label":
        cell.font = Font(name="Calibri", size=11, color="222222")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _THIN_BORDER

    elif style_name == "summary_value":
        color = _RED_FONT if negative else "222222"
        cell.font = Font(name="Calibri", size=11, bold=True, color=color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _THIN_BORDER


def _set_column_widths(ws, num_years):
    ws.column_dimensions["A"].width = 34
    for col in range(2, num_years + 3):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _write_row(ws, row_num, label, values, label_style, data_style, num_format, alt=False):
    label_cell = ws.cell(row=row_num, column=1, value=label)
    _apply_cell_style(label_cell, label_style)
    if alt and label_style == "row_label":
        label_cell.fill = _fill(_ALT_GREY)

    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row_num, column=col_offset + 2)
        if val is None:
            cell.value = "N/D"
            _apply_cell_style(cell, data_style)
        else:
            cell.value = val
            negative = isinstance(val, (int, float)) and val < 0
            _apply_cell_style(cell, data_style, negative=negative)
            cell.number_format = num_format
        if alt and data_style == "data":
            cell.fill = _fill(_ALT_GREY)


# ---------------------------------------------------------------------------
# Constructores de hojas
# ---------------------------------------------------------------------------

def _build_cashflow_sheet(wb, project_data, results):
    ws = wb.active
    ws.title = "Proyección de Flujo de Caja"

    company = project_data["company_name"]
    wacc = project_data["wacc"]
    initial_investment = project_data["initial_investment"]
    num_years = project_data["num_years"]
    year_metrics = results["year_metrics"]

    _set_column_widths(ws, num_years)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 22

    # Fila 1: Título (combinado)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_years + 1)
    title_cell = ws.cell(row=1, column=1, value=f"{company} – Proyección de Flujo de Caja")
    _apply_cell_style(title_cell, "title")

    # Fila 2: Subtítulo
    subtitle = (
        f"WACC: {wacc*100:.2f}%  |  "
        f"Inversión inicial: ${initial_investment:,.2f}  |  "
        f"Años: {num_years}"
    )
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    _apply_cell_style(sub_cell, "subtitle")

    # Fila 4: Encabezados de columnas
    header_label = ws.cell(row=4, column=1, value="Métrica")
    _apply_cell_style(header_label, "col_header")
    for i in range(1, num_years + 1):
        cell = ws.cell(row=4, column=i + 1, value=f"Año {i}")
        _apply_cell_style(cell, "col_header")

    def field_values(key):
        return [m[key] for m in year_metrics]

    # Filas del estado de resultados (5–18)
    rows_def = [
        ("Ingresos",                          "revenue",           "row_label_bold",    "data_section", _FMT_DOLLAR),
        ("Costo de ventas (COGS)",            "cogs",              "row_label",         "data",         _FMT_DOLLAR),
        ("Utilidad bruta",                    "gross_profit",      "row_label_section", "data_section", _FMT_DOLLAR),
        ("  Margen bruto %",                  "gross_margin_pct",  "row_label_margin",  "data_margin",  _FMT_PCT),
        ("Gastos operativos (SG&A)",          "opex",              "row_label",         "data",         _FMT_DOLLAR),
        ("EBITDA",                            "ebitda",            "row_label_section", "data_section", _FMT_DOLLAR),
        ("  Margen EBITDA %",                 "ebitda_margin_pct", "row_label_margin",  "data_margin",  _FMT_PCT),
        ("Depreciación y Amortización (D&A)", "da",                "row_label",         "data",         _FMT_DOLLAR),
        ("EBIT",                              "ebit",              "row_label_bold",    "data",         _FMT_DOLLAR),
        ("Gasto por intereses",               "interest",          "row_label",         "data",         _FMT_DOLLAR),
        ("EBT (Utilidad antes de impuestos)", "ebt",               "row_label_bold",    "data",         _FMT_DOLLAR),
        ("Impuestos",                         "tax",               "row_label",         "data",         _FMT_DOLLAR),
        ("Utilidad neta",                     "net_income",        "row_label_section", "data_section", _FMT_DOLLAR),
        ("  Margen neto %",                   "net_margin_pct",    "row_label_margin",  "data_margin",  _FMT_PCT),
    ]

    alt = False
    for row_offset, (label, key, lbl_style, dat_style, fmt) in enumerate(rows_def):
        row_num = 5 + row_offset
        _write_row(ws, row_num, label, field_values(key), lbl_style, dat_style, fmt, alt=alt)
        if lbl_style == "row_label":
            alt = not alt

    # Fila 19: separador en blanco
    ws.row_dimensions[19].height = 8

    # Filas del flujo de caja libre (20–26)
    fcf_rows = [
        ("Adición D&A",                          "da_addback",                "row_label",         "data",         _FMT_DOLLAR),
        ("CapEx",                                "capex",                     "row_label",         "data",         _FMT_DOLLAR),
        ("Variación en capital de trabajo",      "delta_wc",                  "row_label",         "data",         _FMT_DOLLAR),
        ("Flujo de Caja Libre (FCL)",            "fcf",                       "row_label_section", "data_section", _FMT_DOLLAR),
        ("FCL Descontado",                       "discounted_fcf",            "row_label",         "data",         _FMT_DOLLAR),
        ("FCL Acumulado",                        "cumulative_fcf",            "row_label",         "data",         _FMT_DOLLAR),
        ("FCL Descontado Acumulado",             "cumulative_discounted_fcf", "row_label",         "data",         _FMT_DOLLAR),
    ]

    alt = False
    for row_offset, (label, key, lbl_style, dat_style, fmt) in enumerate(fcf_rows):
        row_num = 20 + row_offset
        _write_row(ws, row_num, label, field_values(key), lbl_style, dat_style, fmt, alt=alt)
        if lbl_style == "row_label":
            alt = not alt

    ws.freeze_panes = "B5"


def _build_summary_sheet(wb, project_data, results):
    ws = wb.create_sheet(title="Resumen Financiero")

    company = project_data["company_name"]
    wacc = project_data["wacc"]
    initial_investment = project_data["initial_investment"]
    num_years = project_data["num_years"]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.row_dimensions[1].height = 26

    # Fila 1: Título
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value=f"Resumen Financiero – {company}")
    _apply_cell_style(title_cell, "title")

    # Fila 2: fecha de generación
    meses = ["enero","febrero","marzo","abril","mayo","junio",
             "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    hoy = date.today()
    fecha_str = f"{hoy.day} de {meses[hoy.month - 1]} de {hoy.year}"
    date_cell = ws.cell(row=2, column=1, value=f"Generado: {fecha_str}")
    _apply_cell_style(date_cell, "subtitle")

    # Fila 4: encabezados
    ws.cell(row=4, column=1, value="Indicador")
    ws.cell(row=4, column=2, value="Valor")
    _apply_cell_style(ws.cell(row=4, column=1), "summary_header")
    _apply_cell_style(ws.cell(row=4, column=2), "summary_header")

    def _summary_row(row, label, value, num_format=None, fill_hex=None):
        lbl = ws.cell(row=row, column=1, value=label)
        _apply_cell_style(lbl, "summary_label")

        negative = isinstance(value, (int, float)) and value < 0
        val_cell = ws.cell(row=row, column=2)
        if value is None:
            val_cell.value = "Fuera del período de proyección"
            _apply_cell_style(val_cell, "summary_value")
        elif isinstance(value, str):
            val_cell.value = value
            _apply_cell_style(val_cell, "summary_value")
        else:
            val_cell.value = value
            _apply_cell_style(val_cell, "summary_value", negative=negative)
            if num_format:
                val_cell.number_format = num_format

        if fill_hex:
            val_cell.fill = _fill(fill_hex)
            lbl.fill = _fill(fill_hex)

    # Fila 5: VPN
    npv = results["npv"]
    npv_fill = _GREEN_FILL if npv >= 0 else _RED_FILL
    _summary_row(5, "Valor Presente Neto (VPN)", npv, _FMT_DOLLAR, npv_fill)

    # Fila 6: TIR
    irr = results["irr"]
    if irr is None:
        _summary_row(6, "Tasa Interna de Retorno (TIR)", "N/D")
    else:
        irr_fill = _GREEN_FILL if irr > wacc else _RED_FILL
        _summary_row(6, "Tasa Interna de Retorno (TIR)", irr, _FMT_PCT, irr_fill)

    # Fila 7: Período de recuperación simple
    sp = results["simple_payback"]
    if sp is None:
        _summary_row(7, "Período de recuperación simple", None)
    else:
        lbl = ws.cell(row=7, column=1, value="Período de recuperación simple (años)")
        _apply_cell_style(lbl, "summary_label")
        val = ws.cell(row=7, column=2, value=sp)
        _apply_cell_style(val, "summary_value")
        val.number_format = _FMT_NUMBER

    # Fila 8: Período de recuperación descontado
    dp = results["discounted_payback"]
    if dp is None:
        _summary_row(8, "Período de recuperación descontado", None)
    else:
        lbl = ws.cell(row=8, column=1, value="Período de recuperación descontado (años)")
        _apply_cell_style(lbl, "summary_label")
        val = ws.cell(row=8, column=2, value=dp)
        _apply_cell_style(val, "summary_value")
        val.number_format = _FMT_NUMBER

    # Filas 10–12: márgenes promedio
    _summary_row(10, "Margen bruto promedio",   results["avg_gross_margin"],   _FMT_PCT)
    _summary_row(11, "Margen EBITDA promedio",  results["avg_ebitda_margin"],  _FMT_PCT)
    _summary_row(12, "Margen neto promedio",    results["avg_net_margin"],     _FMT_PCT)

    # Filas 14–16: referencia de parámetros
    _summary_row(14, "WACC / Tasa de descuento", wacc,                _FMT_PCT)
    _summary_row(15, "Inversión inicial",         initial_investment,  _FMT_DOLLAR)
    _summary_row(16, "Años de proyección",         str(num_years))


# ---------------------------------------------------------------------------
# Punto de entrada público
# ---------------------------------------------------------------------------

def generate_excel(project_data, results, filename):
    """Crea el libro de trabajo, construye las hojas y guarda el archivo."""
    wb = Workbook()
    _build_cashflow_sheet(wb, project_data, results)
    _build_summary_sheet(wb, project_data, results)
    wb.save(filename)
